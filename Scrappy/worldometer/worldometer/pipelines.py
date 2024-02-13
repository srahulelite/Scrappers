# Define your item pipelines here
#
# Don't forget to add your pipeline to the ITEM_PIPELINES setting
# See: https://docs.scrapy.org/en/latest/topics/item-pipeline.html


# useful for handling different item types with a single interface
from itemadapter import ItemAdapter
import logging
import openpyxl
from openpyxl.styles import PatternFill
import os

class WorldometerPipeline:
    def open_spider(self, spider):
        logging.warning("SPIDER OPENED >>>>>>>")

    def close_spider(self, spider):
        logging.warning("SPIDER CLOSED >>>>>>>")

    def process_item(self, item, spider):
        logging.warning(str(type(item))  + ">>>>>>++++")
        return item

class AddToXlsxPipeline:
    def __init__(self):
        self.workbook = None
        self.worksheet = None
        self.current_row_index = 0

    def process_item(self, item, spider):

        values = list(item.values())

        # add header at first row
        if self.current_row_index == 0:
            # get header from item keys
            header = list(item.keys())
            self.worksheet.append(header)
            # After add header, add value below the header
            self.worksheet.append(values)
        else:
            self.worksheet.append(values)
        
        self.current_row_index += 1

        return item

    
    def open_spider(self, spider):
        self.workbook = openpyxl.Workbook()
        self.worksheet = self.workbook.active
    
    def close_spider(self, spider):
        excel_file = '20231107_airmax.xlsx' #file xlsx name
        if(os.path.isfile(excel_file)):
            excel_file = 'new_20231107_airmax.xlsx'
        self.workbook.save(excel_file)
        self.check_diff()

    def check_diff(self):
        df1 = openpyxl.load_workbook("20231107_airmax.xlsx")
        df2 = openpyxl.load_workbook("new_20231107_airmax.xlsx")

        fill_style = PatternFill(start_color="FDD835", end_color="FDD835", fill_type="solid")

        ds1 = df1['Sheet']
        ds2 = df2['Sheet']
        
        for col1 in ds1.iter_cols(max_col=1):
            for cell1 in col1:
                cv = cell1.value #currentValue
                flag_found = False
                for col2 in ds2.iter_cols(max_col=1):
                    for cell2 in col2:
                        if(cell2.value == cv):
                            flag_found = True
                            break
                    if not flag_found:
                        cell1.fill = fill_style
        df1.save("compared.xlsx")
        # df2.save("new_20231107_airmax.xlsx")
